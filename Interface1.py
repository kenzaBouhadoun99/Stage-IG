'''
TODOd:
mettre de la couleur pour le excel
'''
import os
from tkinter import *
from tkinter import filedialog
from tkinter import messagebox
from tkinter import ttk
from tkinter.ttk import Treeview
from matplotlib.rcsetup import validate_color_for_prop_cycle, validate_color_or_auto

import pandas as pd
from PIL import ImageTk, Image
from pyparsing import col

fenetre = Tk()

''' L'interface '''
fenetre.title("MARCIA")
#convertir png en ico 
#fenetre.iconbitmap("Nomimage.ico")
fenetre.geometry("1200x800")

''' Le menu '''
mon_menu = Menu(fenetre)
#creation des onglets
mon_menu.add_cascade(label="File")
mon_menu.add_cascade(label="About")

fenetre.config(menu=mon_menu)

''' Les labels '''
'''Images'''
# Créer les 3 cadres qui représenteront les colonnes
cadre1 = Frame(fenetre, borderwidth=2, relief="solid", width=400, height=400)
cadre2 = Frame(fenetre, borderwidth=2, relief="solid", width=400, height=400)
cadre3 = Frame(fenetre, borderwidth=2, relief="solid", width=400, height=400)
cadre4 = Frame(fenetre, borderwidth=2, relief="solid", width=100, height=100)
cadre5 = Frame(fenetre, borderwidth=2, relief="solid", width=400, height=400)
cadre6 = Frame(fenetre, borderwidth=2, relief="solid", width=100, height=100)
cadre7 = Frame(fenetre, borderwidth=2, relief="solid", width=400, height=400)
# Répartir les colonnes en utilisant la méthode grid dans une ligne et 3 colonnes 
cadre1.grid(row=0, column=0, sticky="nsew")
cadre2.grid(row=0, column=1, sticky="nsew")
cadre3.grid(row=0, column=2, sticky="nsew")
cadre4.grid(row=1, column=0,sticky="nsew")
cadre5.grid(row=2, column=0,sticky="nsew")
cadre6.grid(row=1, column=1,sticky="nsew")
cadre7.grid(row=2, column=1,sticky="nsew")


def calculateur(option):
    pop.destroy()
    if option=='calc':
        print("calc")
    elif option=="ajout":
        print("ajout")
    else:
        global pop2 
        pop2 = Toplevel(fenetre)
        pop2.title("Tools")
        pop2.geometry("350x200")
        #ici je dois ouvrir les diagrammes biplots et ternaires

    pass

def open_settings():
    global pop 
    pop = Toplevel(fenetre)
    pop.title("Tools")
    pop.geometry("350x200")
    pop.config(bg="#909497")

    pop_label =Label(pop,text="VEUILLEZ CHOISIR VOTRE OPTION:")
    pop_label.pack(pady=10)

    my_frame=Frame(pop ,bg="#909497")
    my_frame.pack(pady=5)

    #les boutons de popup 
    calc= Button(my_frame,text="Ouvrir un calculateur d'images",command=lambda: calculateur("calc"))
    calc.grid(row=0 ,column=0)

    ajout= Button(my_frame,text="Ouvrir un diagramme biplot et ternaire",command=lambda: calculateur("ajout"))
    ajout.grid(row=1 ,column=0)

    ouvrir_popup= Button(my_frame,text="Ajouter un élément ou une l’image d’une région spectrale",command=lambda: calculateur("ouvrir_popup"))
    ouvrir_popup.grid(row=2 ,column=0)

# charger l'icône "settings.png" en tant qu'objet PhotoImage et la réduire de moitié
icon_settings = Image.open("images/settings.png")
icon_settings = ImageTk.PhotoImage(icon_settings)

# créer le bouton avec l'icône et sans texte
bouton_parametres = Button(cadre1, image=icon_settings, command=open_settings)
bouton_parametres.pack(side="left", padx=15, pady=15)

my_label = Label(cadre1,text="")

# Ajouter les labels aux cadres correspondants
label2 = Label(cadre1, text="Images")
label2.pack(side="left", padx=65, pady=15)

def save_image():
    global current_image_label
    if current_image_label:
        image = current_image_label.image
        # Convertir la PhotoImage en Image de PIL
        pil_image = ImageTk.getimage(image)
        # Demander à l'utilisateur le nom de fichier pour enregistrer l'image
        fichier = filedialog.asksaveasfilename(title="Enregistrer sous", filetypes=[("Fichier TIFF", "*.tif")], defaultextension=".tif")
        if fichier:
            # Enregistrer l'image au format TIFF 
            pil_image.save(fichier, format="TIFF")




    # sauvegarder l'image avec le nom de fichier sélectionné
# charger l'icône "save.png" en tant qu'objet PhotoImage
icon_save = PhotoImage(file="images/save.png")
icon_save = icon_save.zoom(1)
# créer le bouton avec l'icône et sans texte
bouton_enregistrer = Button(cadre1, image=icon_save, command=save_image)
bouton_enregistrer.pack(side="right", padx=15, pady=15)


# Largeur maximale de l'affichage horizontal
max_width = 300

# Ajouter les noms de fichiers à un widget Label
filename_label = Label(cadre4, text="", font=("Arial", 15))
filename_label.pack()

filenames = [] # une liste vide pour stocker les noms de fichiers
filelist = Listbox(cadre4, font=("Arial", 12)) # une liste Tkinter pour afficher les noms de fichiers
filelist.pack(side=LEFT, fill=BOTH, expand=True) # ajouter l'option expand=True pour remplir complètement le cadre4

filename_str = "" # Initialiser une chaîne vide pour stocker les noms de fichiers formatés horizontalement

# Parcourir chaque élément de la liste et les ajouter au widget Label
for f in filenames:
    # Si la largeur maximale est atteinte, ajouter un retour à la ligne
    if len(filename_str + f) > max_width:
        filename_label.config(text=filename_str)
        filename_str = f + "\n"
    else:
        filename_str += f + " "
filename_label.config(text=filename_str) # Afficher le reste des noms de fichiers



# Configurer le Listbox pour afficher les noms de fichiers horizontalement
filelist.config(width=5, height=5)
filelist.pack(side=LEFT, fill=BOTH, expand=True)


current_image_label = None
def open_file():
    global filenames, current_image_label
    filepath = filedialog.askopenfilename(filetypes=[("Image Files", "*.png;*.jpg;*.jpeg;*.tif")])
    if filepath:
        ext = os.path.splitext(filepath)[1]
        if ext == ".png" or ext == ".jpg" or ext == ".jpeg" or ext == ".tif":
            image = Image.open(filepath)
            max_width = 300 # ou la largeur maximale souhaitée
            max_height = 300 # ou la hauteur maximale souhaitée
            image.thumbnail((max_width, max_height), resample=Image.LANCZOS)
            tk_image = ImageTk.PhotoImage(image)
            # afficher l'image sur un widget Label
            if current_image_label:
                current_image_label.destroy() # Supprimer l'image précédente si elle existe
            current_image_label = Label(cadre5, image=tk_image)
            current_image_label.image = tk_image
            current_image_label.place(x=0, y=0)
            # Ajouter le nom de fichier à la liste des fichiers
            filename = os.path.basename(filepath)
            filenames.append(filename)
            # Mettre à jour la liste des fichiers affichée dans le Listbox
            filelist.delete(0, END)
            for f in filenames:
                filelist.insert(END, f)
            # Mettre à jour la liste des fichiers affichée dans le Listbox avec la fonction update_filelist()
            update_filelist()
        elif ext == ".txt":
            # ouvrir le fichier texte avec le blocs note
            import subprocess
            subprocess.run(["notepad.exe", filepath])
        elif ext == ".raw":
            messagebox.showinfo("Raw Data", "Raw data file selected. Extraction not implemented yet.")
        else:
            messagebox.showerror("Error", "Unsupported file type.")

    

def update_image(event):
    global current_image_label, filenames
    selection = event.widget.curselection()
    if selection:
        filelist.config(exportselection=False)
        # obtenir l'indice du fichier sélectionné
        index = selection[0]
        # obtenir la ligne sélectionnée
        selected_item = event.widget.get(index)
        # extraire le nom de fichier à partir de la ligne sélectionnée
        filename = selected_item.split()[0]
        # construire le chemin d'accès complet à l'image
        script_dir = os.path.dirname(__file__)
        images_dir = os.path.join(script_dir, "images")
        image_path = os.path.join(images_dir, filename)
        # ouvrir l'image et l'afficher dans le Label
        image = Image.open(image_path)
        max_width = 300# ou la largeur maximale souhaitée
        max_height = 300 #ou la hauteur maximale souhaitée
        image.thumbnail((max_width, max_height), resample=Image.Resampling.LANCZOS)
        tk_image = ImageTk.PhotoImage(image)
        if current_image_label:
            current_image_label.destroy() # Supprimer l'image précédente si elle existe
        current_image_label = Label(cadre5, image=tk_image)
        current_image_label.image = tk_image
        current_image_label.place(x=0, y=0)




def update_filelist():
    global file_names_string
    filelist.delete(0, END)
    file_names_string = ""
    for f in filenames:
        # Ajouter le nom de fichier à la chaîne
        file_names_string += f + "    "
        # Si la longueur de la chaîne dépasse la largeur maximale souhaitée
        if filelist.winfo_reqwidth() < len(file_names_string) *7:
            # Ajouter la chaîne à la liste des fichiers et passer à une nouvelle ligne
            filelist.insert(END, file_names_string)
            file_names_string = ""
    # Ajouter la chaîne restante à la liste des fichiers
    filelist.insert(END, file_names_string)
    
    # Afficher les noms de colonnes
    filename_label.config(text="MES IMAGES")
    filename_label.pack(side=TOP)
    
    # Afficher la liste des fichiers
    filelist.pack(side=LEFT, fill=BOTH, expand=1)


# lier la fonction update_image à l'événement de sélection dans le Listbox
filelist.bind("<<ListboxSelect>>", update_image)


# charger l'icône "folder.png" en tant qu'objet PhotoImage et la réduire de moitié
icon_open = PhotoImage(file="images/file.png")
icon_open = icon_open.zoom(1)
# créer le bouton avec l'icône et sans texte
bouton_ouvrir = Button(cadre1, image=icon_open, command=open_file)
bouton_ouvrir.pack(side="right", padx=15, pady=15)



'''Masks'''
'''Save'''
def save_image():
    filename = filedialog.asksaveasfilename(defaultextension=".png")
    # sauvegarder l'image avec le nom de fichier sélectionné
# charger l'icône "folder.png" en tant qu'objet PhotoImage et la réduire de moitié
icon_open4 = PhotoImage(file="images/save.png")
icon_open4 = icon_open4.zoom(1)
# créer le bouton avec l'icône et sans texte
bouton_enregistrer2 = Button(cadre2, image=icon_save, command=save_image)
bouton_enregistrer2.pack(side="right", padx=15, pady=15)

'''Ouvrir'''
import tkinter as tk
from tkinter import ttk
from tkinter import filedialog
import pandas as pd

excel_table = None
filename=""
from os.path import basename

excel_table = None

def open_fileMask():
    global excel_table
    # Ouvrir la boîte de dialogue pour sélectionner un fichier Excel
    filename = filedialog.askopenfilename(filetypes=[("Fichiers Excel", "*.xlsx")])
    print("le nom dans la fonction est", basename(filename))

    if filename:
        # Si un fichier est sélectionné, ouvrir le fichier et afficher son contenu dans le cadre7
        excel_file = pd.read_excel(filename)
        excel_table = ttk.Treeview(cadre7, show="headings")

        # Créer la liste des colonnes en utilisant la méthode 'columns' de 'excel_file'
        columns = list(excel_file.columns)

        excel_table["columns"] = columns
        for column in excel_table["columns"]:
            excel_table.column(column, width=50)

        # Afficher les en-têtes de colonnes
        for column in excel_table["columns"]:
            excel_table.heading(column, text=column)

        # Ajouter les données du fichier
        for row in excel_file.to_numpy():
            excel_table.insert("", "end", values=list(row))

        # Activer l'édition des cellules
        def edit_cell(event):
            # Vérifier que la sélection est sur une cellule (et non une ligne ou une colonne entière)
            region = excel_table.identify_region(event.x, event.y)
            if region == "cell":
                # Obtenir l'index de la ligne et de la colonne sélectionnées
                selection = excel_table.selection()[0]
                row = int(selection[1:])
                column = excel_table.identify_column(event.x)
                if column in columns:
                    col = columns.index(column)
                    # Ouvrir l'éditeur de cellule
                    cell_value = excel_table.item(selection, option='values')[col]
                    excel_table.item(selection, values=[f"{cell_value} (en édition)"])
                    excel_table.update_idletasks()
                    excel_table.edit(selection, col)

        
        excel_table.bind("<Double-1>", edit_cell)

        # Définir la nouvelle configuration de style pour le Treeview
        style = ttk.Style()
        style.configure("Custom.Treeview", rowheight=30)
        excel_table.configure(style="Custom.Treeview")

        excel_table.pack()

        # stocker le nom du fichier sélectionné dans excel_table
        excel_table.filename = filename

def get_selected_filename():
    global excel_table
    if excel_table.filename is None:
        return "Aucun fichier n'a été ouvert."
    else:
        print("le nom du fichier est ",excel_table.filename)
        return excel_table.filename

         



# charger l'icône "folder.png" en tant qu'objet PhotoImage et la réduire de moitié
icon_open2 = PhotoImage(file="images/file.png")
icon_open2 = icon_open2.zoom(1)
# créer le bouton avec l'icône et sans texte

bouton_ouvrir2 = Button(cadre2, image=icon_open2, command=open_fileMask)
bouton_ouvrir2.pack(side="right", padx=15, pady=15)


'''Plus'''
#partie scrolle 
mycanvas =Canvas(cadre6)
mycanvas.pack(side=TOP, fill="both", expand=True)
xscrollbar = Scrollbar(cadre6, orient="horizontal", command=mycanvas.xview)
xscrollbar.pack(side=BOTTOM, fill='x')
mycanvas.configure(xscrollcommand=xscrollbar.set)
mycanvas.bind('<Configure>', lambda e: mycanvas.configure(scrollregion=mycanvas.bbox("all")))
mycanvas.configure(scrollregion=mycanvas.bbox("all"))

myframe=Frame(mycanvas)
myframe.configure(width=mycanvas.winfo_width())
mycanvas.create_window((0,0), window=myframe, anchor="nw")

# initialisation des masques à 1
option_count = 0
checkbutton_list = []
var_list = []
option_count=0
frame7 = None

import openpyxl

def add_checkbutton():
    global option_count, col_count ,frame7
    option_count += 1
    var = IntVar()  # création d'une variable de contrôle pour chaque Checkbutton
    var_list.append(var)
    checkbutton = Checkbutton(myframe, text=f"Masque {option_count}", variable=var)
    checkbutton.grid(row=(option_count-1)//6, column=(option_count-1)%6, padx=2, pady=2)
    checkbutton_list.append(checkbutton)  # ajout du Checkbutton et de sa variable de contrôle à la liste

    # Ajouter une colonne au fichier Excel
    wb = openpyxl.load_workbook(get_selected_filename())
    ws = wb.active
    col_count = ws.max_column + 1  # obtenir le numéro de la nouvelle colonne
    ws.cell(row=1, column=col_count, value=f"Masque {option_count}")  # ajouter l'en-tête de la colonne
    for i in range(2, option_count+2):  # ajouter les cases vides de la nouvelle colonne
        ws.cell(row=i, column=col_count, value="")
    wb.save(get_selected_filename())

    # Mettre à jour le tableau
    update_table()

def update_table():
    global col_count, label_list ,frame7, excel_table
    label_list = []
    wb = openpyxl.load_workbook(get_selected_filename())
    ws = wb.active
    col_count = ws.max_column
    
    # Ajouter les nouvelles colonnes
    new_columns = []
    for j in range(excel_table["columns"], col_count+1):
        header = ws.cell(row=1, column=j).value
        excel_table.column(header, width=50)
        excel_table.heading(header, text=header)
        new_columns.append(header)
    
    # Mettre à jour la liste des colonnes dans excel_table
    excel_table["columns"] += tuple(new_columns)
    
    # Ajouter les nouvelles données
    for row in ws.iter_rows(min_row=2, max_col=col_count, values_only=True):
        excel_table.insert("", "end", values=row)
    
    wb.close()









# charger l'icône "plus.png" en tant qu'objet PhotoImage et la réduire de moitié
icon_plus = PhotoImage(file="images/plus.png")
icon_plus = icon_plus.zoom(1)

bouton_plus = Button(cadre2, image=icon_plus, command=add_checkbutton)
bouton_plus.pack(side="left", padx=2, pady=15)
'''Moin'''
# fonction pour supprimer les Checkbutton sélectionnés
def remove_checkbutton():
    selected = False
    for child in myframe.winfo_children():  # parcours de tous les enfants du cadre contenant les Checkbutton
        if isinstance(child, Checkbutton) and var_list[checkbutton_list.index(child)].get() == 1:
            selected = True
            child.destroy()
    #Vérifier si une boite est déjà sélectionnée ou pas 
    if not selected:
        messagebox.showerror("ATTENTION: ERREUR", "Aucune boîte sélectionnée")

    
# charger l'icône "plus.png" en tant qu'objet PhotoImage et la réduire de moitié
icon_moins = PhotoImage(file="images/moins.png")
icon_moins= icon_moins.zoom(1)

# créer le bouton avec l'icône et sans texte
bouton_moins = Button(cadre2, image=icon_moins, command=remove_checkbutton)
bouton_moins.pack(side="left", padx=2, pady=15)

'''Titre Masks'''
# Ajouter les labels aux cadres correspondants
label2 = Label(cadre2, text="Masks")
label2.pack(side="left", padx=65, pady=15)



'''Classification'''

'''Save'''
def save_image():
    filename = filedialog.asksaveasfilename(defaultextension=".png")
    # sauvegarder l'image avec le nom de fichier sélectionné
# charger l'icône "folder.png" en tant qu'objet PhotoImage et la réduire de moitié
icon_open6 = PhotoImage(file="images/save.png")
icon_open6 = icon_open6.zoom(1)
# créer le bouton avec l'icône et sans texte
bouton_enregistrer6 = Button(cadre3, image=icon_save, command=save_image)
bouton_enregistrer6.pack(side="right", padx=15, pady=15)

'''Titre Classification'''
def open_settings():
    global pop 
    pop = Toplevel(fenetre)
    pop.title("Tools")
    pop.geometry("350x200")
    pop.config(bg="#909497")

    pop_label =Label(pop,text="VEUILLEZ CHOISIR VOTRE OPTION:")
    pop_label.pack(pady=10)

    my_frame=Frame(pop ,bg="#909497")
    my_frame.pack(pady=5)

    #les boutons de popup 
    calc= Button(my_frame,text="Ouvrir un calculateur d'images",command=lambda: calculateur("calc"))
    calc.grid(row=0 ,column=0)

    ajout= Button(my_frame,text="Ouvrir un diagramme biplot et ternaire",command=lambda: calculateur("ajout"))
    ajout.grid(row=1 ,column=0)

    ouvrir_popup= Button(my_frame,text="Ajouter un élément ou une l’image d’une région spectrale",command=lambda: calculateur("ouvrir_popup"))
    ouvrir_popup.grid(row=2 ,column=0)

# charger l'icône "settings.png" en tant qu'objet PhotoImage et la réduire de moitié
icon_settings2 = Image.open("images/settings.png")
icon_settings2 = ImageTk.PhotoImage(icon_settings2)

# créer le bouton avec l'icône et sans texte
bouton_parametres2 = Button(cadre3, image=icon_settings2, command=open_settings)
bouton_parametres2.pack(side="left", padx=15, pady=15)


# charger l'icône "settings.png" en tant qu'objet PhotoImage et la réduire de moitié
icon_update = Image.open("images/MiseAJ.png")
icon_update = ImageTk.PhotoImage(icon_update)

# créer le bouton avec l'icône et sans texte
bouton_update = Button(cadre3, image=icon_update)
bouton_update.pack(side="left", padx=2, pady=15)


# Ajouter les labels aux cadres correspondants
label2 = Label(cadre3, text="Classification")
label2.pack(side="left", padx=25, pady=15)

# Redimensionner les colonnes pour qu'elles s'adaptent à la taille de la fenêtre
fenetre.grid_columnconfigure(0, weight=1)
fenetre.grid_columnconfigure(1, weight=1)
fenetre.grid_columnconfigure(2, weight=1)


mainloop()